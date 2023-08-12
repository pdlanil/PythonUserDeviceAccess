import openpyxl  # Import the openpyxl library for working with Excel files

# Define the filename for the Excel file
user_file = "credentials.xlsx"

# Function to read user credentials from the Excel file
def read_credentials(filename):
    credentials = {}  # Initialize an empty dictionary for credentials
    
    wb = openpyxl.load_workbook(filename)  # Load the Excel workbook
    sheet = wb.active  # Get the active sheet in the workbook
    
    # Iterate through rows starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[0]  # Get username from the first column
        password = row[1]  # Get password from the second column
        credentials[username] = password  # Add username-password pair to dictionary
    
    return credentials  # Return the dictionary of credentials

# Function to write user credentials to the Excel file
def write_credentials(filename, credentials):
    wb = openpyxl.Workbook()  # Create a new workbook
    sheet = wb.active  # Get the active sheet
    
    sheet.append(["Username", "Password"])  # Add header row
    
    # Iterate through credentials dictionary and write to Excel
    for username, password in credentials.items():
        sheet.append([username, password])
    
    wb.save(filename)  # Save the workbook to the given filename

# Function to add a new user to credentials
def add_user(credentials):
    username = input("Enter a new username: ")  # Get new username from user
    password = input("Enter a password: ")  # Get new password from user
    
    credentials[username] = password  # Add new user to credentials dictionary
    write_credentials(user_file, credentials)  # Write updated credentials to Excel
    
    print("User added successfully.")  # Print success message

# Function to view all users in credentials
def view_users(credentials):
    print("List of users:")  # Print header
    for username in credentials.keys():
        print(username)  # Print each username

# Function to delete a user from credentials
def delete_user(credentials):
    username = input("Enter the username to delete: ")  # Get username to delete
    
    if username in credentials:
        del credentials[username]  # Delete user from credentials dictionary
        write_credentials(user_file, credentials)  # Write updated credentials to Excel
        print("User deleted successfully.")  # Print success message
    else:
        print("User not found.")  # Print error message if user not found

# Main function to run the user management interface
def main():
    filename = user_file  # Set the filename
    credentials = read_credentials(filename)  # Load credentials from Excel
    
    while True:
        print("\n1. Add User")
        print("2. View all Users")
        print("3. Delete User")
        print("4. Exit")
        
        choice = input("Enter your choice: ")  # Get user's choice
        
        if choice == "1":
            add_user(credentials)
        elif choice == "2":
            view_users(credentials)
        elif choice == "3":
            delete_user(credentials)
        elif choice == "4":
            break  # Exit the loop if user chooses to exit
        else:
            print("Invalid choice. Please select a valid option.")

# Run the main function if the script is executed directly
if __name__ == "__main__":
    main()
